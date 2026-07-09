"""XLSX fixture ingestion for local option-chain data."""

from __future__ import annotations

from collections.abc import Mapping
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from volsurface.core.schema import NormalizedOptionQuote, OptionType


SUPPORTED_COLUMNS = frozenset(
    {
        "underlying",
        "quote_timestamp",
        "expiry",
        "option_type",
        "strike",
        "bid",
        "ask",
        "last",
        "volume",
        "open_interest",
    }
)
REQUIRED_COLUMNS = frozenset(
    {
        "underlying",
        "expiry",
        "option_type",
        "strike",
        "bid",
        "ask",
    }
)


@dataclass(frozen=True, slots=True)
class XlsxRowError:
    """Row-level XLSX ingestion failure."""

    row_number: int
    code: str
    message: str
    raw_row: Mapping[str, Any] = field(default_factory=dict)


@dataclass(frozen=True, slots=True)
class XlsxIngestionResult:
    """Structured result from reading an XLSX option-chain fixture."""

    quotes: tuple[NormalizedOptionQuote, ...]
    errors: tuple[XlsxRowError, ...]
    source_path: str
    source_name: str
    worksheet_name: str

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)


def read_xlsx_option_chain(
    path: str | Path,
    *,
    worksheet_name: str | None = None,
    source_name: str | None = None,
) -> XlsxIngestionResult:
    """Read a local XLSX fixture into normalized option quotes."""

    source_path = Path(path)
    resolved_source_name = source_name or source_path.name
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook[worksheet_name] if worksheet_name else workbook.active
    resolved_worksheet_name = worksheet.title

    quotes: list[NormalizedOptionQuote] = []
    errors: list[XlsxRowError] = []
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_values = next(rows)
    except StopIteration:
        workbook.close()
        return XlsxIngestionResult(
            quotes=(),
            errors=(
                XlsxRowError(
                    row_number=0,
                    code="EMPTY_XLSX",
                    message="XLSX worksheet does not contain a header row.",
                ),
            ),
            source_path=str(source_path),
            source_name=resolved_source_name,
            worksheet_name=resolved_worksheet_name,
        )

    headers = _normalize_headers(header_values)
    missing_columns = sorted(REQUIRED_COLUMNS.difference(headers))
    if not headers:
        workbook.close()
        return XlsxIngestionResult(
            quotes=(),
            errors=(
                XlsxRowError(
                    row_number=0,
                    code="EMPTY_XLSX",
                    message="XLSX worksheet does not contain a header row.",
                ),
            ),
            source_path=str(source_path),
            source_name=resolved_source_name,
            worksheet_name=resolved_worksheet_name,
        )
    if missing_columns:
        workbook.close()
        return XlsxIngestionResult(
            quotes=(),
            errors=(
                XlsxRowError(
                    row_number=1,
                    code="MISSING_COLUMNS",
                    message=(
                        "XLSX worksheet is missing required columns: "
                        + ", ".join(missing_columns)
                        + "."
                    ),
                ),
            ),
            source_path=str(source_path),
            source_name=resolved_source_name,
            worksheet_name=resolved_worksheet_name,
        )

    try:
        for row_number, values in enumerate(rows, start=2):
            if _is_blank_row(values):
                continue
            raw_row = _row_mapping(headers, values)
            try:
                quotes.append(
                    _parse_quote_row(
                        row=raw_row,
                        row_number=row_number,
                        source_path=source_path,
                        source_name=resolved_source_name,
                    )
                )
            except (TypeError, ValueError) as exc:
                errors.append(
                    XlsxRowError(
                        row_number=row_number,
                        code="ROW_PARSE_ERROR",
                        message=str(exc),
                        raw_row=raw_row,
                    )
                )
    finally:
        workbook.close()

    return XlsxIngestionResult(
        quotes=tuple(quotes),
        errors=tuple(errors),
        source_path=str(source_path),
        source_name=resolved_source_name,
        worksheet_name=resolved_worksheet_name,
    )


def load_xlsx_option_chain(
    path: str | Path,
    *,
    worksheet_name: str | None = None,
    source_name: str | None = None,
) -> XlsxIngestionResult:
    """Alias for callers that prefer loader naming."""

    return read_xlsx_option_chain(
        path,
        worksheet_name=worksheet_name,
        source_name=source_name,
    )


def _parse_quote_row(
    *,
    row: Mapping[str, Any],
    row_number: int,
    source_path: Path,
    source_name: str,
) -> NormalizedOptionQuote:
    return NormalizedOptionQuote(
        underlying=_required_text(row, "underlying"),
        quote_timestamp=_optional_datetime(row, "quote_timestamp"),
        expiry=_required_datetime(row, "expiry"),
        option_type=_parse_option_type(row),
        strike=_required_float(row, "strike"),
        bid=_required_float(row, "bid"),
        ask=_required_float(row, "ask"),
        last=_optional_float(row, "last"),
        volume=_optional_int(row, "volume"),
        open_interest=_optional_int(row, "open_interest"),
        source=source_name,
        metadata={
            "source_path": str(source_path),
            "source_name": source_name,
            "row_number": row_number,
        },
    )


def _normalize_headers(header_values: tuple[Any, ...]) -> tuple[str, ...]:
    headers: list[str] = []
    for value in header_values:
        if value is None:
            headers.append("")
        else:
            headers.append(str(value).strip())
    return tuple(headers)


def _row_mapping(headers: tuple[str, ...], values: tuple[Any, ...]) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        row[header] = values[index] if index < len(values) else None
    return row


def _is_blank_row(values: tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def _parse_option_type(row: Mapping[str, Any]) -> OptionType:
    raw_value = _required_text(row, "option_type").lower()
    try:
        return OptionType(raw_value)
    except ValueError as exc:
        valid_values = ", ".join(member.value for member in OptionType)
        raise ValueError(f"option_type must be one of: {valid_values}") from exc


def _required_text(row: Mapping[str, Any], column: str) -> str:
    value = row.get(column)
    if value is None or not str(value).strip():
        raise ValueError(f"{column} is required")
    return str(value).strip()


def _optional_text(row: Mapping[str, Any], column: str) -> str | None:
    value = row.get(column)
    if value is None or not str(value).strip():
        return None
    return str(value).strip()


def _required_datetime(row: Mapping[str, Any], column: str) -> datetime:
    value = row.get(column)
    if isinstance(value, datetime):
        return value
    return _parse_datetime_value(column, _required_text(row, column))


def _optional_datetime(row: Mapping[str, Any], column: str) -> datetime | None:
    value = row.get(column)
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value
    return _parse_datetime_value(column, _optional_text(row, column) or "")


def _parse_datetime_value(column: str, value: str) -> datetime:
    normalized = value[:-1] + "+00:00" if value.endswith("Z") else value
    try:
        return datetime.fromisoformat(normalized)
    except ValueError as exc:
        raise ValueError(f"{column} must be an ISO-8601 datetime") from exc


def _required_float(row: Mapping[str, Any], column: str) -> float:
    value = row.get(column)
    if value is None or str(value).strip() == "":
        raise ValueError(f"{column} is required")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{column} must be a number") from exc


def _optional_float(row: Mapping[str, Any], column: str) -> float | None:
    value = row.get(column)
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{column} must be a number") from exc


def _optional_int(row: Mapping[str, Any], column: str) -> int | None:
    value = row.get(column)
    if value is None or str(value).strip() == "":
        return None
    try:
        numeric_value = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{column} must be an integer") from exc
    if not numeric_value.is_integer():
        raise ValueError(f"{column} must be an integer")
    return int(numeric_value)


__all__ = [
    "REQUIRED_COLUMNS",
    "SUPPORTED_COLUMNS",
    "XlsxIngestionResult",
    "XlsxRowError",
    "load_xlsx_option_chain",
    "read_xlsx_option_chain",
]
