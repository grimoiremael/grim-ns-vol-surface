from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from volsurface.core.schema import NormalizedOptionQuote, OptionType  # noqa: E402
from volsurface.workbench.xlsx_report import (  # noqa: E402
    NORMALIZED_QUOTE_HEADERS,
    NORMALIZED_QUOTES_SHEET,
    export_normalized_quotes_xlsx,
)


def test_xlsx_report_exports_normalized_quotes_sheet_and_headers(tmp_path: Path) -> None:
    output_path = tmp_path / "normalized_quotes.xlsx"
    quote = NormalizedOptionQuote(
        underlying="SPY",
        quote_timestamp=datetime(2026, 1, 2, 14, 30, tzinfo=timezone.utc),
        expiry=datetime(2026, 2, 20, 21, 0, tzinfo=timezone.utc),
        option_type=OptionType.CALL,
        strike=500.0,
        bid=12.10,
        ask=12.40,
        mid=12.25,
        last=12.20,
        volume=100,
        open_interest=250,
        quote_id="SPY-C-500",
        source="synthetic",
        metadata={"row_number": 2},
    )

    result_path = export_normalized_quotes_xlsx([quote], output_path)

    assert result_path == output_path
    assert output_path.exists()

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        assert NORMALIZED_QUOTES_SHEET in workbook.sheetnames
        worksheet = workbook[NORMALIZED_QUOTES_SHEET]
        headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        row = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]
    finally:
        workbook.close()

    assert tuple(headers) == NORMALIZED_QUOTE_HEADERS
    assert row[0] == "SPY"
    assert row[3] == "call"
    assert row[4] == 500.0
    assert row[5] == 12.10
    assert row[6] == 12.40
    assert '"row_number": 2' in row[-1]


def test_xlsx_report_rejects_non_quote_objects(tmp_path: Path) -> None:
    output_path = tmp_path / "bad.xlsx"

    try:
        export_normalized_quotes_xlsx([object()], output_path)
    except TypeError as exc:
        assert "NormalizedOptionQuote" in str(exc)
    else:
        raise AssertionError("Expected TypeError for non-quote export input")
